from flask import render_template, request, redirect, url_for, flash, current_app, Response
from flask_login import login_required, current_user
from sqlalchemy import distinct, func
from datetime import datetime
from app import db
from app.models import Laporan, SearchPreference
from app.forms import SearchForm, SaveSearchForm
from app.utils import build_search_query, export_search_results, get_search_statistics, format_datetime
from . import bp

# ======================
# DASHBOARD
# ======================
@bp.route("/dashboard")
# @login_required  # DISABLED for development
def dashboard():
    try:
        # Get search parameters
        search_form = SearchForm()
        
        # Populate unit choices dynamically
        units = db.session.query(distinct(Laporan.unit)).filter(Laporan.unit.isnot(None)).all()
        search_form.unit_filter.choices = [('', 'Semua Unit')] + [(unit[0], unit[0]) for unit in units]
        
        # Build query based on search criteria
        if request.args:
            # Populate form with URL parameters
            search_form.search_query.data = request.args.get('search_query', '')
            search_form.unit_filter.data = request.args.get('unit_filter', '')
            search_form.status_filter.data = request.args.get('status_filter', '')
            search_form.jenis_filter.data = request.args.get('jenis_filter', '')
            search_form.pelapor_filter.data = request.args.get('pelapor_filter', '')
            search_form.sort_by.data = request.args.get('sort_by', 'id')
            search_form.sort_order.data = request.args.get('sort_order', 'asc')
            
            # Handle date filters
            if request.args.get('date_from'):
                try:
                    search_form.date_from.data = datetime.strptime(request.args.get('date_from'), '%Y-%m-%d').date()
                except:
                    pass
            
            if request.args.get('date_to'):
                try:
                    search_form.date_to.data = datetime.strptime(request.args.get('date_to'), '%Y-%m-%d').date()
                except:
                    pass
            
            # Build filtered query
            query = build_search_query(request.args)
        else:
            # Default query - sort by ID ascending for sequential order
            query = Laporan.query.order_by(Laporan.id.asc())
        
        # Pagination
        page = request.args.get('page', 1, type=int)
        per_page = 10
        
        laporan = query.paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        # Get search statistics
        search_stats = get_search_statistics(query)
        
        # Get user's saved searches
        saved_searches = []
        # TEMPORARY FIX: Skip saved searches when not authenticated
        # if current_user.is_authenticated:
        #     saved_searches = SearchPreference.query.filter_by(user_id=current_user.id).all()
        
        # Helper for pagination to preserve search queries
        def url_for_page(page):
            args = request.args.copy()
            args['page'] = page
            return url_for('main.dashboard', **args)
            
        return render_template("dashboard_modern.html", 
                             laporan=laporan, 
                             format_datetime=format_datetime,
                             search_form=search_form,
                             search_stats=search_stats,
                             saved_searches=saved_searches,
                             url_for_page=url_for_page)
    except Exception as e:
        current_app.logger.error(f'Dashboard error: {str(e)}', exc_info=True)
        flash('Terjadi kesalahan saat memuat dashboard', 'error')
        # Create empty search form and pagination for error case
        search_form = SearchForm()
        # Create empty pagination object
        from flask_sqlalchemy import Pagination
        empty_pagination = Laporan.query.filter(Laporan.id == -1).paginate(page=1, per_page=10, error_out=False)
        return render_template("dashboard_modern.html", 
                             laporan=empty_pagination, 
                             search_form=search_form,
                             search_stats={'total': 0, 'status_stats': {'pending': 0, 'in_progress': 0, 'resolved': 0}},
                             saved_searches=[],
                             url_for_page=lambda p: url_for('main.dashboard', page=p),
                             format_datetime=format_datetime)

# ======================
# SEARCH & EXPORT ROUTES
# ======================
@bp.route("/search", methods=["GET", "POST"])
# @login_required  # DISABLED for development
def search():
    form = SearchForm()
    
    # Populate unit choices dynamically
    units = db.session.query(distinct(Laporan.unit)).filter(Laporan.unit.isnot(None)).all()
    form.unit_filter.choices = [('', 'Semua Unit')] + [(unit[0], unit[0]) for unit in units]
    
    if form.validate_on_submit():
        # Build query parameters
        params = {}
        if form.search_query.data:
            params['search_query'] = form.search_query.data
        if form.unit_filter.data:
            params['unit_filter'] = form.unit_filter.data
        if form.status_filter.data:
            params['status_filter'] = form.status_filter.data
        if form.jenis_filter.data:
            params['jenis_filter'] = form.jenis_filter.data
        if form.pelapor_filter.data:
            params['pelapor_filter'] = form.pelapor_filter.data
        if form.date_from.data:
            params['date_from'] = form.date_from.data.strftime('%Y-%m-%d')
        if form.date_to.data:
            params['date_to'] = form.date_to.data.strftime('%Y-%m-%d')
        if form.sort_by.data:
            params['sort_by'] = form.sort_by.data
        if form.sort_order.data:
            params['sort_order'] = form.sort_order.data
        
        # Redirect to dashboard with search parameters
        return redirect(url_for('main.dashboard', **params))
    
    return render_template("search.html", form=form)

@bp.route("/export")
# @login_required  # DISABLED for development
def export_results():
    """Export laporan to CSV or Excel"""
    try:
        # Build query based on current search parameters
        query = build_search_query(request.args)
        laporan_list = query.all()
        
        # Export format
        export_format = request.args.get('format', 'csv')
        
        if export_format == 'csv':
            csv_data = export_search_results(laporan_list, 'csv')
            
            filename = f"laporan_simrs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            return Response(
                csv_data,
                mimetype='text/csv',
                headers={'Content-Disposition': f'attachment; filename={filename}'}
            )
        elif export_format == 'excel':
            # Excel export
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Laporan SIMRS"
            
            # Header styling
            header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Headers
            headers = ['ID', 'Unit', 'Pelapor', 'Modul SIMRS', 'Jenis Kesalahan', 
                      'Deskripsi', 'Tanggal Kejadian', 'Status', 'Tanggal Dibuat']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Data rows
            for row_idx, laporan in enumerate(laporan_list, 2):
                ws.cell(row=row_idx, column=1, value=laporan.id)
                ws.cell(row=row_idx, column=2, value=laporan.unit)
                ws.cell(row=row_idx, column=3, value=laporan.pelapor)
                ws.cell(row=row_idx, column=4, value=laporan.modul_simrs or '')
                ws.cell(row=row_idx, column=5, value=laporan.jenis_kesalahan)
                ws.cell(row=row_idx, column=6, value=laporan.deskripsi)
                ws.cell(row=row_idx, column=7, value=laporan.tgl_kejadian.strftime('%Y-%m-%d %H:%M') if laporan.tgl_kejadian else '')
                ws.cell(row=row_idx, column=8, value=laporan.status)
                ws.cell(row=row_idx, column=9, value=laporan.created_at.strftime('%Y-%m-%d %H:%M') if laporan.created_at else '')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"laporan_simrs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': f'attachment; filename={filename}'}
            )
        
        flash('Format export tidak didukung', 'error')
        return redirect(url_for('main.dashboard'))
        
    except Exception as e:
        current_app.logger.error(f'Export error: {str(e)}')
        flash('Terjadi kesalahan saat export data', 'error')
        return redirect(url_for('main.dashboard'))

@bp.route("/save_search", methods=["POST"])
# @login_required  # DISABLED for development
def save_search():
    """Save search preferences"""
    
    form = SaveSearchForm()
    
    if form.validate_on_submit():
        try:
            # Check if search name already exists for this user
            existing = SearchPreference.query.filter_by(
                user_id=current_user.id,
                name=form.name.data
            ).first()
            
            if existing:
                flash('Nama pencarian sudah ada', 'error')
                return redirect(url_for('main.dashboard'))
            
            # Create new search preference
            search_pref = SearchPreference(
                user_id=current_user.id,
                name=form.name.data,
                search_query=request.form.get('search_query'),
                unit_filter=request.form.get('unit_filter'),
                status_filter=request.form.get('status_filter'),
                jenis_filter=request.form.get('jenis_filter'),
                pelapor_filter=request.form.get('pelapor_filter'),
                sort_by=request.form.get('sort_by', 'created_at'),
                sort_order=request.form.get('sort_order', 'desc')
            )
            
            # Handle date filters
            if request.form.get('date_from'):
                try:
                    search_pref.date_from = datetime.strptime(request.form.get('date_from'), '%Y-%m-%d').date()
                except:
                    pass
            
            if request.form.get('date_to'):
                try:
                    search_pref.date_to = datetime.strptime(request.form.get('date_to'), '%Y-%m-%d').date()
                except:
                    pass
            
            db.session.add(search_pref)
            db.session.commit()
            
            flash(f'Pencarian "{form.name.data}" berhasil disimpan', 'success')
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'Save search error: {str(e)}')
            flash('Terjadi kesalahan saat menyimpan pencarian', 'error')
    
    return redirect(url_for('main.dashboard'))

@bp.route("/load_search/<int:search_id>")
# @login_required  # DISABLED for development
def load_search(search_id):
    """Load saved search preferences"""
    
    try:
        search_pref = SearchPreference.query.filter_by(
            id=search_id,
            user_id=current_user.id
        ).first_or_404()
        
        # Build parameters from saved search
        params = {}
        if search_pref.search_query:
            params['search_query'] = search_pref.search_query
        if search_pref.unit_filter:
            params['unit_filter'] = search_pref.unit_filter
        if search_pref.status_filter:
            params['status_filter'] = search_pref.status_filter
        if search_pref.jenis_filter:
            params['jenis_filter'] = search_pref.jenis_filter
        if search_pref.pelapor_filter:
            params['pelapor_filter'] = search_pref.pelapor_filter
        if search_pref.date_from:
            params['date_from'] = search_pref.date_from.strftime('%Y-%m-%d')
        if search_pref.date_to:
            params['date_to'] = search_pref.date_to.strftime('%Y-%m-%d')
        if search_pref.sort_by:
            params['sort_by'] = search_pref.sort_by
        if search_pref.sort_order:
            params['sort_order'] = search_pref.sort_order
        
        return redirect(url_for('main.dashboard', **params))
        
    except Exception as e:
        current_app.logger.error(f'Load search error: {str(e)}')
        flash('Pencarian tidak ditemukan', 'error')
        return redirect(url_for('main.dashboard'))

@bp.route("/delete_search/<int:search_id>", methods=["POST"])
# @login_required  # DISABLED for development
def delete_search(search_id):
    """Delete saved search preferences"""
    
    try:
        search_pref = SearchPreference.query.filter_by(
            id=search_id,
            user_id=current_user.id
        ).first_or_404()
        
        search_name = search_pref.name
        db.session.delete(search_pref)
        db.session.commit()
        
        flash(f'Pencarian "{search_name}" berhasil dihapus', 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Delete search error: {str(e)}')
        flash('Terjadi kesalahan saat menghapus pencarian', 'error')
    
    return redirect(url_for('main.dashboard'))

@bp.route("/statistik")
# @login_required  # DISABLED for development
def statistik():
    try:
        # Get statistics data
        total_laporan = Laporan.query.count()
        pending_count = Laporan.query.filter_by(status='pending').count()
        in_progress_count = Laporan.query.filter_by(status='in_progress').count()
        resolved_count = Laporan.query.filter_by(status='resolved').count()
        
        # Get laporan by jenis kesalahan
        jenis_stats = db.session.query(
            Laporan.jenis_kesalahan, 
            func.count(Laporan.id).label('count')
        ).group_by(Laporan.jenis_kesalahan).all()
        
        # Get recent activity (last 10 reports ordered by ID ascending)
        recent_reports = Laporan.query.order_by(Laporan.id.asc()).limit(10).all()
        
        stats_data = {
            'total_laporan': total_laporan,
            'pending_count': pending_count,
            'in_progress_count': in_progress_count,
            'resolved_count': resolved_count,
            'jenis_stats': jenis_stats,
            'recent_reports': recent_reports
        }
        
        return render_template("statistik_modern.html", stats=stats_data, format_datetime=format_datetime)
        
    except Exception as e:
        current_app.logger.error(f'Error loading statistics: {str(e)}')
        flash('Terjadi kesalahan saat memuat statistik', 'error')
        return redirect(url_for('main.dashboard'))

# ======================
# ERROR HANDLERS
# ======================
@bp.app_errorhandler(404)
def not_found_error(error):
    return render_template('errors/404.html'), 404

@bp.app_errorhandler(500)
def internal_error(error):
    db.session.rollback()
    current_app.logger.error(f'Server Error: {error}')
    return render_template('errors/500.html'), 500

@bp.app_errorhandler(413)
def too_large(error):
    flash('File terlalu besar. Maksimal 16MB.', 'error')
    return redirect(request.url)
